"""
Componente Export Utils - Exportación de datos a PDF y Excel.

Maneja:
- Generación de reportes PDF
- Exportación a Excel
- Botones de descarga
"""

import streamlit as st
import pandas as pd
import io
from datetime import datetime
from typing import Optional, Dict, Any, List

try:
    from fpdf import FPDF
    HAS_FPDF = True
except ImportError:
    HAS_FPDF = False

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class PDFReport(FPDF if HAS_FPDF else object):
    """Clase para generar reportes PDF personalizados."""
    
    def __init__(self):
        if not HAS_FPDF:
            raise ImportError("fpdf2 no está instalado. Ejecuta: pip install fpdf2")
        super().__init__()
        self.set_auto_page_break(auto=True, margin=15)
    
    def header(self):
        """Header del PDF."""
        self.set_font('Arial', 'B', 16)
        self.cell(0, 10, 'Portfolio Selector - Reporte', border=0, ln=True, align='C')
        self.set_font('Arial', 'I', 10)
        self.cell(0, 5, f'Generado: {datetime.now().strftime("%Y-%m-%d %H:%M")}', ln=True, align='C')
        self.ln(10)
    
    def footer(self):
        """Footer del PDF."""
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.cell(0, 10, f'Página {self.page_no()}', align='C')
    
    def add_section_title(self, title: str):
        """Agrega un título de sección."""
        self.set_font('Arial', 'B', 14)
        self.set_fill_color(240, 240, 240)
        self.cell(0, 10, title, ln=True, fill=True)
        self.ln(5)
    
    def add_subsection_title(self, title: str):
        """Agrega un subtítulo."""
        self.set_font('Arial', 'B', 12)
        self.cell(0, 8, title, ln=True)
        self.ln(2)
    
    def add_text(self, text: str):
        """Agrega texto normal."""
        self.set_font('Arial', '', 10)
        self.multi_cell(0, 5, text)
        self.ln(3)
    
    def add_metric(self, label: str, value: str):
        """Agrega una métrica."""
        self.set_font('Arial', 'B', 10)
        self.cell(60, 6, f'{label}:', ln=False)
        self.set_font('Arial', '', 10)
        self.cell(0, 6, value, ln=True)
    
    def add_table(self, df: pd.DataFrame, col_widths: Optional[List[int]] = None):
        """Agrega una tabla al PDF."""
        self.set_font('Arial', 'B', 9)
        
        # Calcular anchos de columna
        if col_widths is None:
            available_width = 190  # A4 width - margins
            col_widths = [available_width // len(df.columns)] * len(df.columns)
        
        # Header
        self.set_fill_color(66, 133, 244)
        self.set_text_color(255, 255, 255)
        for i, col in enumerate(df.columns):
            self.cell(col_widths[i], 7, str(col)[:15], border=1, fill=True, align='C')
        self.ln()
        
        # Data
        self.set_font('Arial', '', 8)
        self.set_text_color(0, 0, 0)
        self.set_fill_color(245, 245, 245)
        
        for idx, row in df.iterrows():
            fill = idx % 2 == 0
            for i, val in enumerate(row):
                self.cell(col_widths[i], 6, str(val)[:15], border=1, fill=fill, align='C')
            self.ln()
        
        self.ln(5)


class ExportManager:
    """Clase para gestionar exportaciones de datos."""
    
    @staticmethod
    def generate_pdf_report(
        perfil: str,
        monto_inversion: float,
        df_portafolio: pd.DataFrame,
        metricas: Dict[str, Any],
        df_metricas_activos: Optional[pd.DataFrame] = None
    ) -> Optional[bytes]:
        """
        Genera un reporte PDF completo.
        
        Args:
            perfil: Nombre del perfil
            monto_inversion: Monto invertido
            df_portafolio: DataFrame con composición
            metricas: Diccionario con métricas de backtest
            df_metricas_activos: DataFrame con métricas por activo
            
        Returns:
            bytes del PDF o None si hay error
        """
        if not HAS_FPDF:
            st.error("La librería fpdf2 no está instalada. Ejecuta: pip install fpdf2")
            return None
        
        try:
            pdf = PDFReport()
            pdf.add_page()
            
            # Información del perfil
            pdf.add_section_title(f'Perfil: {perfil.title()}')
            pdf.add_metric('Monto de Inversión', f'${monto_inversion:,.2f} USD')
            pdf.add_metric('Fecha de Generación', datetime.now().strftime('%Y-%m-%d %H:%M'))
            pdf.add_metric('Número de Activos', str(len(df_portafolio)))
            pdf.ln(5)
            
            # Métricas de rendimiento
            pdf.add_section_title('Métricas de Rendimiento')
            
            if metricas:
                pdf.add_metric('Retorno Total', f"{metricas.get('retorno_total', 0)*100:.2f}%")
                pdf.add_metric('CAGR', f"{metricas.get('cagr', 0)*100:.2f}%")
                pdf.add_metric('Volatilidad', f"{metricas.get('volatilidad', 0)*100:.2f}%")
                pdf.add_metric('Sharpe Ratio', f"{metricas.get('sharpe', 0):.2f}")
                pdf.add_metric('Max Drawdown', f"{metricas.get('max_drawdown', 0)*100:.2f}%")
                pdf.add_metric('Sortino Ratio', f"{metricas.get('sortino', 0):.2f}")
            else:
                pdf.add_text('Métricas no disponibles')
            
            pdf.ln(5)
            
            # Composición del portafolio
            pdf.add_section_title('Composición del Portafolio')
            
            df_tabla = df_portafolio.copy()
            df_tabla['monto'] = df_tabla['peso'] * monto_inversion
            df_tabla = df_tabla[['ticker', 'segmento', 'peso', 'monto']]
            df_tabla.columns = ['Ticker', 'Seg.', 'Peso', 'Monto']
            df_tabla['Peso'] = df_tabla['Peso'].apply(lambda x: f'{x*100:.1f}%')
            df_tabla['Monto'] = df_tabla['Monto'].apply(lambda x: f'${x:,.0f}')
            
            pdf.add_table(df_tabla, col_widths=[40, 25, 40, 60])
            
            # Métricas por activo (si están disponibles)
            if df_metricas_activos is not None and not df_metricas_activos.empty:
                pdf.add_page()
                pdf.add_section_title('Métricas por Activo')
                
                df_activos = df_metricas_activos[['ticker', 'retorno_total', 'volatilidad', 'sharpe']].copy()
                df_activos.columns = ['Ticker', 'Retorno', 'Vol.', 'Sharpe']
                df_activos['Retorno'] = df_activos['Retorno'].apply(lambda x: f'{x*100:.1f}%')
                df_activos['Vol.'] = df_activos['Vol.'].apply(lambda x: f'{x*100:.1f}%')
                df_activos['Sharpe'] = df_activos['Sharpe'].apply(lambda x: f'{x:.2f}')
                
                pdf.add_table(df_activos, col_widths=[40, 50, 50, 50])
            
            # Disclaimer
            pdf.ln(10)
            pdf.set_font('Arial', 'I', 8)
            pdf.multi_cell(0, 4, 
                'Disclaimer: Este reporte es unicamente informativo. '
                'Los rendimientos pasados no garantizan rendimientos futuros. '
                'Consulte con un asesor financiero antes de tomar decisiones de inversion.'
            )
            
            # fpdf2: output() ya devuelve bytes directamente
            return bytes(pdf.output())
            
        except Exception as e:
            st.error(f"Error generando PDF: {str(e)}")
            return None
    
    @staticmethod
    def generate_excel_report(
        perfil: str,
        monto_inversion: float,
        df_portafolio: pd.DataFrame,
        metricas: Dict[str, Any],
        df_metricas_activos: Optional[pd.DataFrame] = None,
        df_equity: Optional[pd.DataFrame] = None
    ) -> Optional[bytes]:
        """
        Genera un reporte Excel con múltiples hojas.
        
        Args:
            perfil: Nombre del perfil
            monto_inversion: Monto invertido
            df_portafolio: DataFrame con composición
            metricas: Diccionario con métricas
            df_metricas_activos: DataFrame con métricas por activo
            df_equity: DataFrame con equity curve
            
        Returns:
            bytes del Excel o None si hay error
        """
        if not HAS_OPENPYXL:
            st.error("La librería openpyxl no está instalada. Ejecuta: pip install openpyxl")
            return None
        
        try:
            output = io.BytesIO()
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Hoja 1: Resumen
                df_resumen = pd.DataFrame({
                    'Parámetro': [
                        'Perfil',
                        'Monto de Inversión',
                        'Número de Activos',
                        'Fecha de Generación',
                        '',
                        'Retorno Total',
                        'CAGR',
                        'Volatilidad',
                        'Sharpe Ratio',
                        'Max Drawdown',
                        'Sortino Ratio',
                    ],
                    'Valor': [
                        perfil.title(),
                        f'${monto_inversion:,.2f}',
                        len(df_portafolio),
                        datetime.now().strftime('%Y-%m-%d %H:%M'),
                        '',
                        f"{metricas.get('retorno_total', 0)*100:.2f}%" if metricas else 'N/A',
                        f"{metricas.get('cagr', 0)*100:.2f}%" if metricas else 'N/A',
                        f"{metricas.get('volatilidad', 0)*100:.2f}%" if metricas else 'N/A',
                        f"{metricas.get('sharpe', 0):.2f}" if metricas else 'N/A',
                        f"{metricas.get('max_drawdown', 0)*100:.2f}%" if metricas else 'N/A',
                        f"{metricas.get('sortino', 0):.2f}" if metricas else 'N/A',
                    ]
                })
                df_resumen.to_excel(writer, sheet_name='Resumen', index=False)
                
                # Hoja 2: Composición del portafolio
                df_comp = df_portafolio.copy()
                df_comp['monto_usd'] = df_comp['peso'] * monto_inversion
                df_comp['peso_pct'] = df_comp['peso'] * 100
                df_comp = df_comp[['ticker', 'segmento', 'peso_pct', 'monto_usd']]
                df_comp.columns = ['Ticker', 'Segmento', 'Peso (%)', 'Monto (USD)']
                df_comp.to_excel(writer, sheet_name='Composicion', index=False)
                
                # Hoja 3: Métricas por activo
                if df_metricas_activos is not None and not df_metricas_activos.empty:
                    df_metricas_activos.to_excel(writer, sheet_name='Metricas_Activos', index=False)
                
                # Hoja 4: Equity curve
                if df_equity is not None and not df_equity.empty:
                    df_equity.to_excel(writer, sheet_name='Equity_Curve', index=True)
            
            return output.getvalue()
            
        except Exception as e:
            st.error(f"Error generando Excel: {str(e)}")
            return None
    
    @staticmethod
    def generate_csv_portfolio(df_portafolio: pd.DataFrame, monto_inversion: float) -> bytes:
        """
        Genera un CSV simple del portafolio.
        
        Args:
            df_portafolio: DataFrame con composición
            monto_inversion: Monto invertido
            
        Returns:
            bytes del CSV
        """
        df = df_portafolio.copy()
        df['monto_usd'] = df['peso'] * monto_inversion
        df['peso_pct'] = df['peso'] * 100
        
        return df.to_csv(index=False).encode('utf-8')


def render_export_buttons(
    perfil: str,
    monto_inversion: float,
    df_portafolio: pd.DataFrame,
    metricas: Dict[str, Any],
    df_metricas_activos: Optional[pd.DataFrame] = None,
    df_equity: Optional[pd.DataFrame] = None
):
    """
    Renderiza los botones de exportación.
    
    Args:
        perfil: Nombre del perfil
        monto_inversion: Monto invertido
        df_portafolio: DataFrame con composición
        metricas: Diccionario con métricas
        df_metricas_activos: DataFrame con métricas por activo
        df_equity: DataFrame con equity curve
    """
    st.subheader("Exportar Datos")
    
    col1, col2, col3 = st.columns(3)
    
    fecha = datetime.now().strftime('%Y%m%d')
    
    with col1:
        # Botón CSV
        csv_data = ExportManager.generate_csv_portfolio(df_portafolio, monto_inversion)
        st.download_button(
            label="Descargar CSV",
            data=csv_data,
            file_name=f"portafolio_{perfil}_{fecha}.csv",
            mime="text/csv",
            help="Descarga la composición del portafolio en formato CSV"
        )
    
    with col2:
        # Botón Excel
        if HAS_OPENPYXL:
            excel_data = ExportManager.generate_excel_report(
                perfil, monto_inversion, df_portafolio, 
                metricas, df_metricas_activos, df_equity
            )
            if excel_data:
                st.download_button(
                    label="Descargar Excel",
                    data=excel_data,
                    file_name=f"reporte_{perfil}_{fecha}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    help="Descarga el reporte completo en formato Excel"
                )
        else:
            st.button("Excel (no disponible)", disabled=True)
            st.caption("Instalar: pip install openpyxl")
    
    with col3:
        # Botón PDF
        if HAS_FPDF:
            pdf_data = ExportManager.generate_pdf_report(
                perfil, monto_inversion, df_portafolio,
                metricas, df_metricas_activos
            )
            if pdf_data:
                st.download_button(
                    label="Descargar PDF",
                    data=pdf_data,
                    file_name=f"reporte_{perfil}_{fecha}.pdf",
                    mime="application/pdf",
                    help="Descarga el reporte en formato PDF"
                )
        else:
            st.button("PDF (no disponible)", disabled=True)
            st.caption("Instalar: pip install fpdf2")


def render_quick_export(df_portafolio: pd.DataFrame, monto_inversion: float, perfil: str):
    """
    Renderiza botón de exportación rápida (solo CSV).
    
    Args:
        df_portafolio: DataFrame con composición
        monto_inversion: Monto invertido
        perfil: Nombre del perfil
    """
    csv_data = ExportManager.generate_csv_portfolio(df_portafolio, monto_inversion)
    fecha = datetime.now().strftime('%Y%m%d')
    
    st.download_button(
        label="Exportar",
        data=csv_data,
        file_name=f"portafolio_{perfil}_{fecha}.csv",
        mime="text/csv",
        use_container_width=True
    )
